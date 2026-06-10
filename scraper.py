"""
MenuRadar Scraper v4 — با ScraperAPI
"""

import pandas as pd
import requests
from bs4 import BeautifulSoup
import re, json, os, time, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# تنظیمات
# ─────────────────────────────────────────────

SCRAPER_API_KEY = os.environ.get("SCRAPER_API_KEY", "")
MATCH_THRESHOLD = 0.20

def make_url(target_url):
    """همه درخواست‌ها از ScraperAPI رد می‌شن"""
    if SCRAPER_API_KEY:
        return f"http://api.scraperapi.com?api_key={SCRAPER_API_KEY}&url={target_url}&country_code=ir"
    return target_url

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
    "Accept-Language": "fa,en;q=0.9",
}

# mapping: شیت رقبا → کانسپت منوی من
CONCEPT_MAP = {
    "کافه ":                 "لیبرو",
    "کباب ایرانی ":          "اسپیئو رویال",
    "تایلندی و ایتالیایی":   "روماتای",
    "پیترا و سوخاری ":       "چیکن فکتوری ",
    "برگر":                  "برگر فکتوری",
    "غذای ایرانی ":          "اسپیئو اکسپرس",
    "شیکبار":                "لفو",
}

SYNONYMS = {
    "فرایز": ["سیب زمینی", "فرنچ فرایز", "فری", "fries"],
    "برگر": ["همبرگر", "burger"],
    "مرغ": ["چیکن", "chicken"],
    "سوخاری": ["فراید", "crispy", "fried"],
    "پاستا": ["ماکارونی", "پنه", "اسپاگتی", "pasta"],
    "پیتزا": ["پیزا", "pizza"],
    "قهوه": ["coffee", "لاته", "latte", "اسپرسو", "espresso", "کاپوچینو"],
    "شیک": ["میلک شیک", "milkshake"],
    "کباب": ["کبابی", "kabab"],
}

# ─────────────────────────────────────────────
# خواندن فایل‌ها
# ─────────────────────────────────────────────

def load_my_menu():
    df = pd.read_excel("data/my_menu.xlsx", sheet_name="لیست نهایی")
    df.columns = df.columns.str.strip().str.replace("\u200c", "")
    df = df[df["نام فارسي"].notna()]
    df = df[df["فی واحد  با ارزش افزوده - ریال"].notna()]
    df = df[df["فی واحد  با ارزش افزوده - ریال"] > 100000]
    if "نوع کالا" in df.columns:
        df = df[~df["نوع کالا"].isin(["ظروف", "بسته بندی", "متفرقه"])]
    print(f"✓ My menu: {len(df)} items across {df['کانسپت'].nunique()} concepts")
    return df

def load_competitors():
    xl = pd.ExcelFile("data/competitors.xlsx")
    result = {}
    for sheet in xl.sheet_names:
        df = pd.read_excel("data/competitors.xlsx", sheet_name=sheet)
        df.columns = df.columns.str.strip()
        df = df[df["منبع"].notna()]
        df = df[df["منبع"].str.startswith("http", na=False)]
        result[sheet] = df
        print(f"  '{sheet}': {len(df)} competitors")
    return result

# ─────────────────────────────────────────────
# Scraping
# ─────────────────────────────────────────────

def clean_price(text):
    digits = re.sub(r"[^\d]", "", str(text))
    return int(digits) if digits else None

def to_toman(p):
    if not p:
        return None
    p = int(p)
    return p // 10 if p > 5_000_000 else p

def fetch(url, timeout=30):
    """fetch با ScraperAPI"""
    try:
        proxy_url = make_url(url)
        r = requests.get(proxy_url, headers=HEADERS, timeout=timeout)
        print(f"    HTTP {r.status_code} (len={len(r.text)})")
        return r if r.status_code == 200 else None
    except Exception as e:
        print(f"    fetch error: {e}")
        return None

def scrape_snappfood(url, brand):
    items = []
    # روش ۱: API
    slug = re.search(r"menu/([^/?#]+)", url)
    if slug:
        for api in [
            f"https://snappfood.ir/restaurant/api/v2/restaurant/details?restaurantId={slug.group(1)}",
            f"https://snappfood.ir/mobile/v2/restaurant/details?restaurantId={slug.group(1)}&optionalClient=WEBSITE",
        ]:
            try:
                r = requests.get(make_url(api), headers=HEADERS, timeout=20)
                if r.status_code == 200:
                    data = r.json()
                    # پیدا کردن products در هر سطحی
                    txt = json.dumps(data)
                    products_raw = re.findall(r'"title"\s*:\s*"([^"]+)"[^}]*"price"\s*:\s*(\d+)', txt)
                    for name, price in products_raw[:200]:
                        p = int(price)
                        if p > 1000:
                            items.append({"brand": brand, "item": name, "price_toman": to_toman(p)})
                    if items:
                        return items
            except:
                pass

    # روش ۲: scrape صفحه
    r = fetch(url)
    if not r:
        return items
    soup = BeautifulSoup(r.text, "html.parser")

    # JSON در script
    for script in soup.find_all("script"):
        txt = script.string or ""
        if "price" not in txt.lower() or len(txt) < 50:
            continue
        pairs = re.findall(r'"(?:title|name)"\s*:\s*"([^"]{2,60})"[^}]{0,200}"price"\s*:\s*(\d{4,})', txt)
        for name, price in pairs[:200]:
            items.append({"brand": brand, "item": name, "price_toman": to_toman(int(price))})
        if items:
            break

    return items

def scrape_menew(url, brand):
    items = []
    r = fetch(url)
    if not r:
        return items
    soup = BeautifulSoup(r.text, "html.parser")

    for card in soup.select(".product-card, .menu-item, .food-card, [class*='product'], [class*='item']"):
        name_el = card.select_one("h1,h2,h3,h4,.name,.title,[class*='name'],[class*='title']")
        price_el = card.select_one(".price,[class*='price']")
        if name_el and price_el:
            p = clean_price(price_el.get_text())
            if p and p > 1000:
                items.append({"brand": brand, "item": name_el.get_text(strip=True)[:80],
                              "price_toman": to_toman(p)})
    return items

def scrape_generic(url, brand):
    items = []
    r = fetch(url)
    if not r:
        return items
    soup = BeautifulSoup(r.text, "html.parser")

    # ۱. JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if not isinstance(data, dict):
                continue
            menu = data.get("hasMenu") or data.get("menu", {})
            if isinstance(menu, dict):
                for section in menu.get("hasMenuSection", []):
                    for mi in section.get("hasMenuItem", []):
                        n = mi.get("name", "")
                        p = mi.get("offers", {}).get("price", 0)
                        if n and p:
                            pv = int(float(p))
                            items.append({"brand": brand, "item": n, "price_toman": to_toman(pv)})
        except:
            pass
    if items:
        return items

    # ۲. JSON در script
    for script in soup.find_all("script"):
        txt = script.string or ""
        if len(txt) < 100 or "price" not in txt.lower():
            continue
        for pat in [
            r'"(?:title|name)"\s*:\s*"([^"]{2,60})"[^}]{0,300}"(?:price|cost)"\s*:\s*(\d{4,})',
            r'"(?:price|cost)"\s*:\s*(\d{4,})[^}]{0,300}"(?:title|name)"\s*:\s*"([^"]{2,60})"',
        ]:
            matches = re.findall(pat, txt)
            for m in matches[:150]:
                if len(m) == 2:
                    name, price = (m[0], m[1]) if not m[0].isdigit() else (m[1], m[0])
                    items.append({"brand": brand, "item": name[:80],
                                  "price_toman": to_toman(int(price))})
        if items:
            break
    if items:
        return items

    # ۳. HTML selectors
    selector_groups = [
        ("[class*='menu-item']", "[class*='title'],[class*='name'],h3,h4", "[class*='price']"),
        ("[class*='product']",   "[class*='title'],[class*='name'],h3,h4", "[class*='price']"),
        ("[class*='food']",      "[class*='title'],[class*='name'],h3,h4", "[class*='price']"),
        ("li",                   "[class*='name'],[class*='title']",       "[class*='price']"),
    ]
    for cont, name_s, price_s in selector_groups:
        cards = soup.select(cont)
        for card in cards[:300]:
            ne = card.select_one(name_s)
            pe = card.select_one(price_s)
            if ne and pe:
                p = clean_price(pe.get_text())
                if p and 5000 < p < 999_999_999:
                    items.append({"brand": brand, "item": ne.get_text(strip=True)[:80],
                                  "price_toman": to_toman(p)})
        if items:
            break
    if items:
        return items

    # ۴. Generic fallback
    price_re = re.compile(r"\b(\d[\d,]{3,})\b")
    seen = set()
    for tag in soup.find_all(["div","li","tr","p","span"], limit=3000):
        if tag.find(["div","table","ul","ol"]):
            continue
        text = tag.get_text(separator=" ", strip=True)
        if not (5 < len(text) < 200):
            continue
        prices = [int(m.group(1).replace(",","")) for m in price_re.finditer(text)]
        valid = [p for p in prices if 5000 < p < 999_999_999]
        if not valid:
            continue
        name = price_re.sub("", text).strip()
        name = re.sub(r"[,،٬؛:\-/\\|٪%]", " ", name)
        name = re.sub(r"\s+", " ", name).strip()[:80]
        if len(name) < 3 or name in seen:
            continue
        seen.add(name)
        items.append({"brand": brand, "item": name, "price_toman": to_toman(valid[0])})

    # deduplicate
    seen_items = set()
    unique = []
    for it in items:
        k = it["item"].strip().lower()
        if k and k not in seen_items and it["price_toman"] and it["price_toman"] > 500:
            seen_items.add(k)
            unique.append(it)
    return unique

def scrape_brand(url, brand):
    print(f"  → {brand}")
    url = url.strip()
    if not url.startswith("http"):
        return []

    if any(d in url for d in ["snappfood.ir", "snapp-store.com", "snapp.ir"]):
        items = scrape_snappfood(url, brand)
    elif "menew.ir" in url:
        items = scrape_menew(url, brand)
    else:
        items = scrape_generic(url, brand)

    print(f"    ✓ {len(items)} items scraped")
    time.sleep(2)
    return items

# ─────────────────────────────────────────────
# Matching
# ─────────────────────────────────────────────

def normalize(text):
    text = str(text).lower().strip()
    text = text.replace("ي","ی").replace("ك","ک").replace("‌"," ")
    text = re.sub(r"\s+"," ", text)
    for canon, variants in SYNONYMS.items():
        for v in variants:
            text = text.replace(v.lower(), canon)
    return text

def match_score(a, b):
    wa = set(normalize(a).split())
    wb = set(normalize(b).split())
    if not wa or not wb:
        return 0.0
    inter = wa & wb
    if not inter:
        return 0.0
    return (len(inter)/len(wa|wb) + len(inter)/min(len(wa),len(wb))) / 2

def best_match(my_item, comp_items):
    best, score = None, 0.0
    for ci in comp_items:
        s = match_score(my_item, ci["item"])
        if s > score:
            score, best = s, ci
    return (best, score) if score >= MATCH_THRESHOLD else (None, 0.0)

# ─────────────────────────────────────────────
# Excel Builder
# ─────────────────────────────────────────────

def build_excel(my_df, comp_by_concept):
    wb = Workbook()
    wb.remove(wb.active)

    HDR  = PatternFill("solid", start_color="1F3864")
    SUB  = PatternFill("solid", start_color="2E75B6")
    GRN  = PatternFill("solid", start_color="E2EFDA")
    RED  = PatternFill("solid", start_color="FFDAD9")
    YLW  = PatternFill("solid", start_color="FFF2CC")
    ALT  = PatternFill("solid", start_color="EBF3FB")
    WHT  = PatternFill("solid", start_color="FFFFFF")

    hf = Font(bold=True, color="FFFFFF", name="Tahoma", size=11)
    sf = Font(bold=True, color="FFFFFF", name="Tahoma", size=9)
    nf = Font(name="Tahoma", size=9)
    t  = Side(style="thin", color="BFBFBF")
    bd = Border(left=t, right=t, top=t, bottom=t)
    cc = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rc = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    COLS  = ["دسته‌بندی","نام آیتم (من)","قیمت من (تومان)","نام آیتم رقیب","قیمت رقیب (تومان)","نام رقیب","اختلاف (تومان)","وضعیت"]
    WIDTHS= [20, 38, 16, 38, 16, 22, 16, 14]
    all_rows = []

    for sheet_name, my_concept in CONCEPT_MAP.items():
        comp_items = comp_by_concept.get(sheet_name, [])
        subset = my_df[my_df["کانسپت"].str.strip() == my_concept.strip()].copy()
        if subset.empty:
            continue

        ws = wb.create_sheet(title=my_concept.strip()[:31])
        ws.sheet_view.rightToLeft = True

        ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
        ws["A1"] = f"📊 {my_concept.strip()}  |  آیتم‌های رقبا: {len(comp_items)}"
        ws["A1"].font = Font(bold=True, color="FFFFFF", name="Tahoma", size=12)
        ws["A1"].fill = HDR; ws["A1"].alignment = cc
        ws.row_dimensions[1].height = 30

        for c,h in enumerate(COLS,1):
            cl = ws.cell(row=2,column=c,value=h)
            cl.font=sf; cl.fill=SUB; cl.alignment=cc; cl.border=bd
        ws.row_dimensions[2].height = 22

        row = 3
        matched_count = 0
        for _,r in subset.iterrows():
            my_name  = str(r["نام فارسي"]).strip()
            price_r  = r.get("فی واحد  با ارزش افزوده - ریال", 0)
            my_price = int(float(price_r)) // 10 if price_r else 0
            cat      = str(r.get("نوع کالا","—")).strip()

            bm, sc = best_match(my_name, comp_items)

            if bm:
                matched_count += 1
                cp   = bm["price_toman"] or 0
                diff = my_price - cp
                if diff < -5000:   status,rfill = "🟢 ارزان‌تر", GRN
                elif diff > 5000:  status,rfill = "🔴 گران‌تر",  RED
                else:              status,rfill = "🟡 مشابه",    YLW
                vals = [cat, my_name, my_price, bm["item"], cp, bm["brand"], diff, status]
                all_rows.append({"کانسپت":my_concept.strip(),"دسته":cat,
                                  "آیتم_من":my_name,"قیمت_من":my_price,
                                  "آیتم_رقیب":bm["item"],"قیمت_رقیب":cp,
                                  "رقیب":bm["brand"],"اختلاف":diff,"وضعیت":status})
            else:
                rfill = ALT if row%2==0 else WHT
                vals  = [cat, my_name, my_price, "—","—","—","—","⚪ بدون تطابق"]

            for c,v in enumerate(vals,1):
                cl = ws.cell(row=row,column=c,value=v)
                cl.font=nf; cl.fill=rfill; cl.alignment=rc; cl.border=bd
            row += 1

        for i,w in enumerate(WIDTHS,1):
            ws.column_dimensions[get_column_letter(i)].width = w
        print(f"  ✓ {my_concept.strip()}: {len(subset)} items, {matched_count} matched")

    # شیت خلاصه کل
    if all_rows:
        ws0 = wb.create_sheet("📋 خلاصه کل", 0)
        ws0.sheet_view.rightToLeft = True
        C0 = ["کانسپت","دسته","آیتم من","قیمت من (تومان)","آیتم رقیب","قیمت رقیب (تومان)","رقیب","اختلاف (تومان)","وضعیت"]
        W0 = [16,16,38,16,38,16,22,16,14]

        ws0.merge_cells(f"A1:{get_column_letter(len(C0))}1")
        ws0["A1"] = f"📋 خلاصه کل MenuRadar — {datetime.date.today()}"
        ws0["A1"].font = Font(bold=True, color="FFFFFF", name="Tahoma", size=13)
        ws0["A1"].fill = HDR; ws0["A1"].alignment = cc
        ws0.row_dimensions[1].height = 32

        for c,h in enumerate(C0,1):
            cl = ws0.cell(row=2,column=c,value=h)
            cl.font=sf; cl.fill=SUB; cl.alignment=cc; cl.border=bd
        ws0.row_dimensions[2].height = 22

        for r,item in enumerate(all_rows,3):
            if   item["وضعیت"]=="🟢 ارزان‌تر": rf=GRN
            elif item["وضعیت"]=="🔴 گران‌تر":  rf=RED
            elif item["وضعیت"]=="🟡 مشابه":    rf=YLW
            else: rf = ALT if r%2==0 else WHT
            vals=[item["کانسپت"],item["دسته"],item["آیتم_من"],item["قیمت_من"],
                  item["آیتم_رقیب"],item["قیمت_رقیب"],item["رقیب"],item["اختلاف"],item["وضعیت"]]
            for c,v in enumerate(vals,1):
                cl=ws0.cell(row=r,column=c,value=v)
                cl.font=nf; cl.fill=rf; cl.alignment=rc; cl.border=bd
        for i,w in enumerate(W0,1):
            ws0.column_dimensions[get_column_letter(i)].width = w

    return wb

# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    print("🚀 MenuRadar v4 starting...")
    if SCRAPER_API_KEY:
        print("✓ ScraperAPI enabled")
    else:
        print("⚠ No ScraperAPI key — direct requests (may be blocked)")

    print("\n📂 Loading my menu...")
    my_df = load_my_menu()

    print("\n📂 Loading competitors...")
    comp_sheets = load_competitors()

    print("\n🌐 Scraping competitors...")
    comp_by_concept = {}
    for sheet_name, df in comp_sheets.items():
        print(f"\n📍 {sheet_name}")
        items = []
        for _, row in df.iterrows():
            brand = str(row.get("نام برند","")).strip()
            url   = str(row.get("منبع","")).strip()
            if url.startswith("http"):
                items.extend(scrape_brand(url, brand))
        comp_by_concept[sheet_name] = items
        print(f"  Total: {len(items)} items")

    print("\n📊 Building Excel...")
    wb = build_excel(my_df, comp_by_concept)

    os.makedirs("reports", exist_ok=True)
    out = f"reports/MenuRadar_{datetime.date.today()}.xlsx"
    wb.save(out)
    print(f"\n✅ Done: {out}")
    return out

if __name__ == "__main__":
    main()
