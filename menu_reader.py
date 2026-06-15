# -*- coding: utf-8 -*-
"""menu_reader.py — v9: Playwright + auto-tab + structured parser + regex"""
import datetime, os, re, json, html, time, csv, urllib3, shutil
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings()
HEADERS = {"User-Agent": "Mozilla/5.0 Chrome/120"}
TIMEOUT = 30
MAX_PRICE = 100_000_000
PLAYWRIGHT_TIMEOUT = 60_000
CACHE_FILE = "extract_cache.json"
SCREENSHOTS_DIR = "reports/screenshots"
CHROME_PATH = shutil.which("chromium-browser") or shutil.which("chromium")
IMG_CACHE_KEYS = {"بابل تیله": "img:https://bubbleteale.com/wp-content/uploads/2025/09/menu-site.webp"}


def _cache_get(key):
    try:
        c = json.load(open(CACHE_FILE, encoding="utf-8"))
        e = c.get(key)
        return e["items"] if e else None
    except:
        return None


def scrape_snappfood(url, brand):
    m = re.search(r"-r-([0-9a-zA-Z]+)", url)
    if not m:
        return [], "vendor code پیدا نشد", None
    api = ("https://snappfood.ir/mobile/v2/restaurant/details/dynamic"
           "?optionalClient=WEBSITE&client=WEBSITE&deviceType=WEBSITE"
           "&appVersion=8.1.1&UDID=&vendorCode=" + m.group(1))
    try:
        data = requests.get(api, headers=HEADERS, timeout=TIMEOUT, verify=False).json()
    except Exception as e:
        return [], "API error: {}".format(e), None
    items, seen = [], set()
    cur_cat = ["?"]

    def walk(o):
        if isinstance(o, dict):
            t, p = o.get("title"), o.get("price")
            if t and isinstance(p, (int, float)) and 0 < p < MAX_PRICE and str(t) not in seen:
                seen.add(str(t))
                items.append({"item": str(t).strip(), "price_toman": int(p),
                              "brand": brand, "category": cur_cat[0]})
            for k, v in o.items():
                if k in ("category", "categoryName") and isinstance(v, str):
                    cur_cat[0] = v
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)
    walk(data)
    return items, "", None


def scrape_menew(url, brand):
    m = re.search(r"https?://([^./]+)\.menew\.ir", url)
    if not m:
        return [], "ساب‌دامین menew نامعتبر", None
    link = m.group(1)
    H = {"Origin": "https://{}.menew.ir".format(link),
         "Referer": "https://{}.menew.ir/".format(link)}
    H.update(HEADERS)
    API = "https://citadel.menew.ir/api"
    Q1 = "query Q($link: String!) { entity(link: $link) { id } }"
    Q2 = ("query Q($id: UUID!) { entity(id: $id) { menus { categories "
          "{ label status items { name shopItem { shopItemPrice { price originalPrice } } } } } } }")
    try:
        r = requests.post(API, json={"query": Q1, "variables": {"link": link}},
                          timeout=TIMEOUT, verify=False, headers=H)
        eid = ((r.json().get("data") or {}).get("entity") or {}).get("id")
        if not eid:
            return [], "entity پیدا نشد", None
        r2 = requests.post(API, json={"query": Q2, "variables": {"id": eid}},
                           timeout=60, verify=False, headers=H)
        data = (r2.json().get("data") or {}).get("entity") or {}
    except Exception as e:
        return [], "menew error: {}".format(e), None
    items, seen = [], set()
    for mn in data.get("menus") or []:
        for cat in mn.get("categories") or []:
            cl = re.sub(r"https?://\S+", "", (cat.get("label") or "").strip()).strip()
            for it in cat.get("items") or []:
                name = (it.get("name") or "").strip()
                sp = ((it.get("shopItem") or {}).get("shopItemPrice") or {})
                price = sp.get("price") or sp.get("originalPrice") or 0
                if name and price and 0 < price < MAX_PRICE and name not in seen:
                    seen.add(name)
                    items.append({"item": name, "price_toman": int(price),
                                  "brand": brand, "category": cl})
    return items, "", None


def scrape_image_cache(brand):
    key = IMG_CACHE_KEYS.get(brand.strip())
    if not key:
        return [], "کش تصویری وجود ندارد", None
    c = _cache_get(key)
    if not c:
        return [], "کش تصویری خالی", None
    return [dict(it, brand=brand) for it in c], "از کش Vision قبلی", None


PRICE_RX = re.compile(r"(\d{1,3}(?:[,،]\d{3})+|\d{4,9})\s*(?:تومان|ریال|toman)?", re.I)


def _clean_html(t):
    t = re.sub(r"<script[^>]*>.*?</script>", " ", t, flags=re.S | re.I)
    t = re.sub(r"<style[^>]*>.*?</style>", " ", t, flags=re.S | re.I)
    t = re.sub(r"<[^>]+>", " ", t)
    t = html.unescape(t)
    return re.sub(r"\s+", " ", t)


def scrape_structured(html_text, brand):
    """پارسر ساختاری برای سایت‌هایی که h2/h3/h4 + strong/b دارن (مثل خروس/Delino)"""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return []
    try:
        soup = BeautifulSoup(html_text, 'html.parser')
    except Exception:
        return []
    items = []
    seen = set()
    for h in soup.find_all(['h2', 'h3', 'h4']):
        name = h.get_text(strip=True)
        if not name or len(name) < 2 or len(name) > 80 or name in seen:
            continue
        parent = h.find_parent()
        for _ in range(5):
            if parent is None:
                break
            price_tag = None
            for tag in parent.find_all(['strong', 'b']):
                txt = tag.get_text()
                if 'تومان' in txt or 'ریال' in txt:
                    price_tag = tag
                    break
            if price_tag:
                mp = re.search(r'(\d{1,3}(?:[,،]\d{3})+|\d{4,9})', price_tag.get_text())
                if mp:
                    try:
                        price = int(mp.group(1).replace(',', '').replace('،', ''))
                        if 5000 <= price <= MAX_PRICE:
                            seen.add(name)
                            items.append({'item': name, 'price_toman': price,
                                          'brand': brand, 'category': ''})
                            break
                    except:
                        pass
            parent = parent.find_parent()
    return items


def scrape_regex(html_text, brand):
    txt = _clean_html(html_text)
    items, seen = [], set()
    for m in PRICE_RX.finditer(txt):
        chunk = txt[max(0, m.start() - 150):m.start()].strip()
        parts = chunk.split()
        if len(parts) < 2:
            continue
        name = None
        for take in (4, 5, 6, 7, 3, 8):
            cand = " ".join(parts[-take:]).strip(" .،,:-—|/")
            if 6 <= len(cand) <= 100 and not cand.replace(" ", "").isdigit():
                cand = re.sub(r"^(از|تا|قیمت|هزینه|مبلغ|تومان|ریال)\s+", "", cand)
                if len(cand) >= 6:
                    name = cand
                    break
        if not name:
            continue
        try:
            price = int(m.group(1).replace(",", "").replace("،", ""))
        except:
            continue
        if price < 5000 or price > MAX_PRICE * 10:
            continue
        if price >= 1000000 and price % 10 == 0:
            price //= 10
        if price > MAX_PRICE:
            continue
        if name not in seen:
            seen.add(name)
            items.append({"item": name, "price_toman": price, "brand": brand, "category": ""})
        if len(items) >= 500:
            break
    return items


_PW = {"p": None, "browser": None}


def _get_browser():
    if _PW["browser"]:
        return _PW["browser"]
    from playwright.sync_api import sync_playwright
    p = sync_playwright().start()
    b = p.chromium.launch(headless=True, executable_path=CHROME_PATH,
                          args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
                                "--disable-blink-features=AutomationControlled"])
    _PW["p"], _PW["browser"] = p, b
    return b


def _close_browser():
    if _PW["browser"]:
        try:
            _PW["browser"].close()
        except:
            pass
    if _PW["p"]:
        try:
            _PW["p"].stop()
        except:
            pass


def _safe_filename(name):
    return re.sub(r"[^\w\u0600-\u06FF]+", "_", name).strip("_")


def _take_screenshot(page, brand):
    os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
    path = os.path.join(SCREENSHOTS_DIR, _safe_filename(brand) + ".png")
    try:
        page.screenshot(path=path, full_page=True)
        return path
    except:
        return None


EXPAND_JS = """
async () => {
  const sleep = ms => new Promise(r => setTimeout(r, ms));
  let clicked = 0;
  document.querySelectorAll('[style*="display: none"], [style*="display:none"], .hidden, .collapse, [hidden]').forEach(el => {
    try {
      el.style.display = 'block';
      el.removeAttribute('hidden');
      el.classList.remove('hidden', 'collapse');
    } catch(e) {}
  });
  const selectors = [
    'button[role="tab"]', '[role="tab"]', '.tab', '.tab-link', '.nav-tabs a',
    '.category-tab', '.menu-tab', '[class*="category-btn"]', '[class*="categoryBtn"]',
    '[class*="filter-btn"]', '[class*="filterBtn"]',
    '.accordion-button', '.accordion-header button', '.collapse-toggle',
    'a[data-toggle="tab"]', 'a[data-bs-toggle="tab"]',
    '[data-tab]', '[data-category]',
    '.swiper-slide button', '.MuiTab-root',
    'nav button', '.menu-categories button', '.categories button',
    'li[role="tab"]', 'div[role="tab"]'
  ];
  for (const sel of selectors) {
    const els = document.querySelectorAll(sel);
    for (const el of els) {
      try {
        el.click();
        clicked++;
        await sleep(150);
      } catch(e) {}
    }
  }
  for (let i = 0; i < 5; i++) {
    window.scrollTo(0, document.body.scrollHeight * (i+1) / 5);
    await sleep(400);
  }
  window.scrollTo(0, document.body.scrollHeight);
  await sleep(800);
  return clicked;
}
"""


def scrape_playwright(url, brand):
    if not CHROME_PATH:
        return [], "chromium نصب نیست", None
    screenshot_path = None
    last_err = ""
    urls_to_try = [url]
    if "://www." not in url:
        urls_to_try.append(url.replace("://", "://www.", 1))
    elif "://www." in url:
        urls_to_try.append(url.replace("://www.", "://", 1))
    for try_url in urls_to_try:
        try:
            b = _get_browser()
            ctx = b.new_context(user_agent="Mozilla/5.0 Chrome/120",
                                viewport={"width": 1366, "height": 768},
                                ignore_https_errors=True)
            page = ctx.new_page()
            try:
                page.goto(try_url, timeout=PLAYWRIGHT_TIMEOUT, wait_until="networkidle")
            except:
                try:
                    page.goto(try_url, timeout=PLAYWRIGHT_TIMEOUT, wait_until="domcontentloaded")
                except Exception as e:
                    last_err = str(e)[:60]
                    ctx.close()
                    continue
            time.sleep(2)
            try:
                clicked = page.evaluate(EXPAND_JS)
            except Exception:
                clicked = 0
            time.sleep(2)
            html_text = page.content()
            if len(html_text) < 1000:
                ctx.close()
                last_err = "صفحه خیلی کوچک"
                continue

            # اول structured (h3+strong) - کیفیت بالاتر
            struct_items = scrape_structured(html_text, brand)
            if struct_items and len(struct_items) >= 5:
                ctx.close()
                return struct_items, "PW+structured ({} آیتم)".format(len(struct_items)), None

            # بعد regex
            items = scrape_regex(html_text, brand)
            if items:
                ctx.close()
                msg = "PW+regex" + (" (تب‌ها: {})".format(clicked) if clicked else "")
                return items, msg, None

            # اگه هیچی نبود، اسکرین‌شات
            screenshot_path = _take_screenshot(page, brand)
            ctx.close()
            break
        except Exception as e:
            last_err = "PW error: {}".format(str(e)[:60])
    if screenshot_path:
        return [], "regex خالی — عکس: {}".format(os.path.basename(screenshot_path)), screenshot_path
    return [], last_err or "خطای نامعلوم", None


def scrape_brand(url, brand):
    bn = brand.strip()
    if bn in IMG_CACHE_KEYS:
        return scrape_image_cache(brand)
    u = (url or "").lower()
    if "menew.ir" in u:
        return scrape_menew(url, brand)
    if "snappfood.ir" in u or "snapp-store.com" in u:
        return scrape_snappfood(url, brand)
    if not url:
        return [], "بدون URL", None
    return scrape_playwright(url, brand)


def load_competitors(path="data/competitors.xlsx"):
    wb = load_workbook(path, read_only=True)
    data = {}
    for sh in wb.sheetnames:
        rows = list(wb[sh].iter_rows(values_only=True))[1:]
        entries = []
        for r in rows:
            if not r or not r[0]:
                continue
            brand = str(r[0]).strip()
            url = next((str(c).strip() for c in r if c and str(c).strip().startswith("http")), None)
            entries.append((brand, url))
        data[sh] = entries
    return data


def build_csv(results, path):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["کانسپت", "برند", "دسته", "آیتم", "قیمت (تومان)"])
        for sheet, entries in results.items():
            for brand, url, items, err, _ in entries:
                for it in items:
                    w.writerow([sheet.strip(), brand, it.get("category", ""),
                                it["item"], it["price_toman"]])


def build_excel(results, path):
    wb = Workbook()
    wb.remove(wb.active)
    HDR = PatternFill("solid", start_color="1F3864")
    SUB = PatternFill("solid", start_color="2E75B6")
    OK = PatternFill("solid", start_color="E2EFDA")
    BAD = PatternFill("solid", start_color="FFDAD9")
    SHOT = PatternFill("solid", start_color="FFF2CC")
    ALT = PatternFill("solid", start_color="EBF3FB")
    WHT = PatternFill("solid", start_color="FFFFFF")
    hf = Font(bold=True, color="FFFFFF", name="Tahoma", size=11)
    sf = Font(bold=True, color="FFFFFF", name="Tahoma", size=9)
    nf = Font(name="Tahoma", size=9)
    t = Side(style="thin", color="BFBFBF")
    bd = Border(left=t, right=t, top=t, bottom=t)
    cc = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rc = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws = wb.create_sheet("Status")
    ws.sheet_view.rightToLeft = True
    ws.merge_cells("A1:D1")
    ws["A1"] = "وضعیت — {}".format(datetime.date.today())
    ws["A1"].font = hf
    ws["A1"].fill = HDR
    ws["A1"].alignment = cc
    for c, h in enumerate(["کانسپت", "برند", "تعداد آیتم", "منبع/خطا"], 1):
        cl = ws.cell(row=2, column=c, value=h)
        cl.font = sf
        cl.fill = SUB
        cl.alignment = cc
        cl.border = bd
    row = 3
    total = 0
    live = 0
    shots = 0
    for sheet, entries in results.items():
        for brand, url, items, err, shot in entries:
            n = len(items)
            total += n
            if n > 0:
                live += 1
            if shot:
                shots += 1
            fill = OK if n > 0 else (SHOT if shot else BAD)
            for c, v in enumerate([sheet.strip(), brand, n, err], 1):
                cl = ws.cell(row=row, column=c, value=v)
                cl.font = nf
                cl.fill = fill
                cl.alignment = rc
                cl.border = bd
            row += 1
    for i, w in enumerate([30, 25, 12, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    safe = {"کافه": "Cafe", "(رویال)کباب ایرانی": "Royal",
            "تایلندی و ایتالیایی(روماتای)": "Romatay",
            "پیترا و سوخاری (چیکن فکتوری)": "ChickenFactory",
            "(برگرفکتوری)برگر": "BurgerFactory",
            "غذای ایرانی ( اکسپرس)": "Express", "شیکبار(لفو)": "Lefoo"}
    for sheet, entries in results.items():
        title = safe.get(sheet.strip(), sheet.strip())[:31] or "Sheet"
        wsc = wb.create_sheet(title=title)
        wsc.sheet_view.rightToLeft = True
        wsc.merge_cells("A1:D1")
        total_c = sum(len(it) for _, _, it, _, _ in entries)
        wsc["A1"] = "{} — جمع {} آیتم".format(sheet.strip(), total_c)
        wsc["A1"].font = hf
        wsc["A1"].fill = HDR
        wsc["A1"].alignment = cc
        for c, h in enumerate(["برند", "دسته", "آیتم", "قیمت (تومان)"], 1):
            cl = wsc.cell(row=2, column=c, value=h)
            cl.font = sf
            cl.fill = SUB
            cl.alignment = cc
            cl.border = bd
        r = 3
        for brand, url, items, err, shot in entries:
            wsc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            cl = wsc.cell(row=r, column=1,
                          value="— {} ({} آیتم) {}".format(brand, len(items), err))
            cl.font = Font(bold=True, name="Tahoma", size=9)
            cl.fill = OK if items else (SHOT if shot else BAD)
            cl.alignment = rc
            cl.border = bd
            r += 1
            for it in items:
                vals = [brand, it.get("category", ""), it["item"], it["price_toman"]]
                fill = ALT if r % 2 == 0 else WHT
                for c, v in enumerate(vals, 1):
                    cl = wsc.cell(row=r, column=c, value=v)
                    cl.font = nf
                    cl.fill = fill
                    cl.alignment = rc
                    cl.border = bd
                r += 1
        for i, w in enumerate([20, 22, 45, 16], 1):
            wsc.column_dimensions[get_column_letter(i)].width = w
        wsc.freeze_panes = "A3"
    wb.save(path)
    return live, total, shots


def main():
    print("=" * 60)
    print("MenuRadar v9 — Playwright + structured + regex")
    print("=" * 60)
    print("Chromium:", CHROME_PATH)
    print()
    competitors = load_competitors()
    results = {}
    try:
        for sheet, entries in competitors.items():
            print("\n### {} ({} برند)".format(sheet.strip(), len(entries)))
            out = []
            for brand, url in entries:
                items, err, shot = scrape_brand(url or "", brand)
                badge = "✔" if items else ("📷" if shot else "✗")
                print("  {} {:25s} → {:4d} آیتم  {}".format(badge, brand[:25], len(items), err[:60]))
                out.append((brand, url, items, err, shot))
            results[sheet] = out
    finally:
        _close_browser()
    os.makedirs("reports", exist_ok=True)
    today = datetime.date.today()
    xlsx = "reports/RawMenus_{}.xlsx".format(today)
    csvp = "reports/menus_{}.csv".format(today)
    live, total, shots = build_excel(results, xlsx)
    build_csv(results, csvp)
    print("\n" + "=" * 60)
    print("✅ {}".format(xlsx))
    print("✅ {}".format(csvp))
    print("📊 {} برند زنده | جمع {} آیتم | {} اسکرین‌شات".format(live, total, shots))
    print("=" * 60)


if __name__ == "__main__":
    main()
