import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import re
import json
import os
from urllib.parse import urlparse

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36"
}

SNAPPFOOD_DOMAINS = ["snappfood.ir", "snapp-store.com", "superapp.snappfood.ir"]
MENEW_DOMAINS = ["menew.ir"]

# ---------------------------------------------------------------------------
# Concept mapping: competitor sheet name → my concept name
# ---------------------------------------------------------------------------
CONCEPT_MAP = {
    "کافه": "لیبرو",
    "کباب ایرانی": "اسپیئو رویال",
    "تایلندی و ایتالیایی": "روماتای",
    "پیترا و سوخاری": "چیکن فکتوری",
    "برگر": "برگر فکتوری",
    "غذای ایرانی": "اسپیئو اکسپرس",
    "شیکبار": "لفو",
}

# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------

def load_my_menu(path="data/my_menu.xlsx"):
    df = pd.read_excel(path, sheet_name="لیست نهایی")
    df = df[df["نوع کالا"].notna()]
    # exclude packaging/equipment rows
    df = df[~df["نوع کالا"].isin(["ظروف"])]
    df = df[df["فی واحد  با ارزش افزوده - ریال"] > 100]
    return df

def load_competitors(path="data/competitors.xlsx"):
    xl = pd.read_excel(path, sheet_name=None)
    result = {}
    for sheet_name, df in xl.items():
        if sheet_name not in CONCEPT_MAP:
            continue
        if df.empty or df.shape[1] < 2:
            continue
        df.columns = ["نام_برند", "URL"]
        df = df.dropna(subset=["URL"])
        df = df[df["URL"].str.startswith("http", na=False)]
        result[sheet_name] = df
    return result

# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------

def get_domain(url):
    return urlparse(url).netloc.replace("www.", "")

def is_snappfood(url):
    return any(d in url for d in SNAPPFOOD_DOMAINS)

def is_menew(url):
    return any(d in url for d in MENEW_DOMAINS)

def parse_price(text):
    text = re.sub(r"[^\d]", "", str(text))
    return int(text) if text else None

def scrape_snappfood(url, brand_name):
    """Scrape snappfood menu via their API"""
    items = []
    try:
        # Extract restaurant slug from URL
        match = re.search(r"menu/([^/?]+)", url)
        if not match:
            return items
        slug = match.group(1)
        api_url = f"https://snappfood.ir/restaurant/api/v1/details/menu?restaurantId={slug}"
        # Try direct page scrape instead
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return items
        soup = BeautifulSoup(r.text, "html.parser")
        # snappfood loads via JS - try to find JSON in script tags
        for script in soup.find_all("script"):
            if script.string and "products" in str(script.string):
                try:
                    data_match = re.search(r'"products"\s*:\s*(\[.*?\])', script.string, re.DOTALL)
                    if data_match:
                        products = json.loads(data_match.group(1))
                        for p in products:
                            name = p.get("title") or p.get("name", "")
                            price = p.get("price") or p.get("basePrice", 0)
                            if name and price:
                                items.append({
                                    "نام_رقیب": brand_name,
                                    "نام_آیتم_رقیب": name,
                                    "قیمت_رقیب_ریال": int(price) * 10 if int(price) < 1000000 else int(price),
                                    "قیمت_رقیب_تومان": int(price) if int(price) < 1000000 else int(price) // 10,
                                })
                except:
                    pass
        # Fallback: parse HTML product cards
        if not items:
            for card in soup.select("[class*='product'], [class*='item'], [class*='food']"):
                name_el = card.select_one("[class*='title'], [class*='name']")
                price_el = card.select_one("[class*='price']")
                if name_el and price_el:
                    price = parse_price(price_el.get_text())
                    if price and price > 1000:
                        items.append({
                            "نام_رقیب": brand_name,
                            "نام_آیتم_رقیب": name_el.get_text(strip=True),
                            "قیمت_رقیب_ریال": price * 10 if price < 1000000 else price,
                            "قیمت_رقیب_تومان": price if price < 1000000 else price // 10,
                        })
    except Exception as e:
        print(f"  ⚠ snappfood error for {brand_name}: {e}")
    return items

def scrape_menew(url, brand_name):
    items = []
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return items
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select(".product-card, .menu-item, [class*='product'], [class*='item']"):
            name_el = card.select_one(".product-name, .item-name, h3, h4, [class*='name'], [class*='title']")
            price_el = card.select_one(".price, [class*='price']")
            if name_el and price_el:
                price = parse_price(price_el.get_text())
                if price and price > 1000:
                    items.append({
                        "نام_رقیب": brand_name,
                        "نام_آیتم_رقیب": name_el.get_text(strip=True),
                        "قیمت_رقیب_ریال": price * 10 if price < 1000000 else price,
                        "قیمت_رقیب_تومان": price if price < 1000000 else price // 10,
                    })
    except Exception as e:
        print(f"  ⚠ menew error for {brand_name}: {e}")
    return items

def scrape_generic(url, brand_name):
    items = []
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return items
        soup = BeautifulSoup(r.text, "html.parser")
        # Try JSON-LD
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string)
                if isinstance(data, dict) and data.get("@type") == "Restaurant":
                    menu = data.get("hasMenu", {})
                    for section in menu.get("hasMenuSection", []):
                        for item in section.get("hasMenuItem", []):
                            name = item.get("name", "")
                            price = item.get("offers", {}).get("price", 0)
                            if name and price:
                                price = int(float(price))
                                items.append({
                                    "نام_رقیب": brand_name,
                                    "نام_آیتم_رقیب": name,
                                    "قیمت_رقیب_ریال": price * 10 if price < 1000000 else price,
                                    "قیمت_رقیب_تومان": price if price < 1000000 else price // 10,
                                })
            except:
                pass
        # Fallback generic selectors
        if not items:
            price_pattern = re.compile(r"[\d,]{3,}")
            candidates = []
            for el in soup.find_all(["div", "li", "tr", "article"]):
                text = el.get_text(separator=" ", strip=True)
                prices = price_pattern.findall(text.replace(",", ""))
                valid_prices = [int(p) for p in prices if 10000 < int(p) < 500000000]
                if valid_prices and len(text) < 300:
                    # Try to extract name (text before price)
                    name_part = re.sub(r"[\d,،٬]+", "", text).strip()
                    name_part = re.sub(r"\s+", " ", name_part).strip()
                    if 2 < len(name_part) < 80:
                        p = valid_prices[0]
                        candidates.append({
                            "نام_رقیب": brand_name,
                            "نام_آیتم_رقیب": name_part[:80],
                            "قیمت_رقیب_ریال": p * 10 if p < 1000000 else p,
                            "قیمت_رقیب_تومان": p if p < 1000000 else p // 10,
                        })
            # deduplicate by name
            seen = set()
            for c in candidates:
                k = c["نام_آیتم_رقیب"]
                if k not in seen:
                    seen.add(k)
                    items.append(c)
    except Exception as e:
        print(f"  ⚠ generic error for {brand_name}: {e}")
    return items

def scrape_brand(url, brand_name):
    print(f"  → Scraping {brand_name} ...")
    if is_snappfood(url):
        items = scrape_snappfood(url, brand_name)
    elif is_menew(url):
        items = scrape_menew(url, brand_name)
    else:
        items = scrape_generic(url, brand_name)
    print(f"    ✓ {len(items)} items found")
    time.sleep(2)
    return items

# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

SYNONYMS = {
    "فرایز": ["سیب زمینی", "فرنچ فرایز", "فرایز", "فری"],
    "برگر": ["همبرگر", "برگر"],
    "چیز": ["پنیر"],
    "چیکن": ["مرغ", "chicken"],
    "کبابی": ["کباب"],
    "سوخاری": ["فراید", "fried", "crispy"],
    "پاستا": ["ماکارونی", "پنه", "اسپاگتی"],
    "پیتزا": ["پیزا"],
    "لاته": ["latte", "لته"],
    "اسپرسو": ["espresso"],
    "کاپوچینو": ["cappuccino", "کاپو"],
}

def normalize(text):
    text = str(text).lower().strip()
    text = re.sub(r"[‌\s]+", " ", text)
    for canonical, variants in SYNONYMS.items():
        for v in variants:
            text = text.replace(v.lower(), canonical.lower())
    return text

def match_score(my_item, competitor_item):
    my_words = set(normalize(my_item).split())
    comp_words = set(normalize(competitor_item).split())
    if not my_words or not comp_words:
        return 0
    intersection = my_words & comp_words
    union = my_words | comp_words
    # Jaccard similarity
    jaccard = len(intersection) / len(union)
    # Bonus for longer common subsequence
    common_ratio = len(intersection) / min(len(my_words), len(comp_words))
    return (jaccard + common_ratio) / 2

MATCH_THRESHOLD = 0.25

def find_best_match(my_item_name, competitor_items):
    best_score = 0
    best_match = None
    for item in competitor_items:
        score = match_score(my_item_name, item["نام_آیتم_رقیب"])
        if score > best_score:
            best_score = score
            best_match = item
    if best_score >= MATCH_THRESHOLD:
        return best_match, best_score
    return None, 0

# ---------------------------------------------------------------------------
# Build output Excel
# ---------------------------------------------------------------------------

def build_output(my_menu_df, competitors_data, concept_map):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    HEADER_FILL = PatternFill("solid", start_color="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")
    SUBHEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    ALT_FILL = PatternFill("solid", start_color="EBF3FB")
    NORMAL_FONT = Font(name="Arial", size=9)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    summary_rows = []

    for sheet_name, my_concept in concept_map.items():
        if sheet_name not in competitors_data:
            continue

        comp_df = competitors_data[sheet_name]
        my_items = my_menu_df[my_menu_df["کانسپت"] == my_concept].copy()
        my_items = my_items[["نام فارسي", "نوع کالا", "فی واحد  با ارزش افزوده - ریال"]].copy()
        my_items.columns = ["نام_آیتم_من", "دسته_بندی", "قیمت_من_ریال"]
        my_items["قیمت_من_تومان"] = (my_items["قیمت_من_ریال"] / 10).astype(int)

        if my_items.empty:
            continue

        # Collect all competitor items for this concept
        all_comp_items = competitors_data.get(f"_items_{sheet_name}", [])

        ws = wb.create_sheet(title=my_concept[:31])

        # Title row
        ws.merge_cells("A1:H1")
        ws["A1"] = f"📊 مقایسه منو — {my_concept}  |  تعداد رقبا: {len(comp_df)}"
        ws["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        ws["A1"].fill = PatternFill("solid", start_color="1F3864")
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 28

        # Column headers (row 2)
        headers = [
            "دسته‌بندی", "نام آیتم (من)",
            "قیمت من (تومان)", "نام آیتم رقیب",
            "قیمت رقیب (تومان)", "نام رقیب",
            "اختلاف قیمت (تومان)", "وضعیت"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = CENTER
            cell.border = border
        ws.row_dimensions[2].height = 22

        row = 3
        matched_count = 0
        for _, my_row in my_items.iterrows():
            best_match, score = find_best_match(my_row["نام_آیتم_من"], all_comp_items)
            fill = ALT_FILL if row % 2 == 0 else PatternFill("solid", start_color="FFFFFF")

            my_price = int(my_row["قیمت_من_تومان"])
            if best_match:
                matched_count += 1
                comp_price = int(best_match["قیمت_رقیب_تومان"])
                diff = my_price - comp_price
                status = "🟢 ارزان‌تر" if diff < -5000 else ("🔴 گران‌تر" if diff > 5000 else "🟡 مشابه")
                vals = [
                    my_row["دسته_بندی"],
                    my_row["نام_آیتم_من"],
                    my_price,
                    best_match["نام_آیتم_رقیب"],
                    comp_price,
                    best_match["نام_رقیب"],
                    diff,
                    status,
                ]
                summary_rows.append({
                    "کانسپت": my_concept,
                    "آیتم_من": my_row["نام_آیتم_من"],
                    "قیمت_من": my_price,
                    "آیتم_رقیب": best_match["نام_آیتم_رقیب"],
                    "قیمت_رقیب": comp_price,
                    "رقیب": best_match["نام_رقیب"],
                    "اختلاف": diff,
                    "وضعیت": status,
                })
            else:
                vals = [
                    my_row["دسته_بندی"],
                    my_row["نام_آیتم_من"],
                    my_price,
                    "—", "—", "—", "—", "⚪ بدون تطابق"
                ]

            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = NORMAL_FONT
                cell.fill = fill
                cell.alignment = RIGHT if col in [3, 5, 7] else Alignment(horizontal="right", vertical="center", wrap_text=True)
                cell.border = border
            row += 1

        # Column widths
        col_widths = [18, 35, 18, 35, 18, 22, 20, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.sheet_view.rightToLeft = True

    # Summary sheet
    if summary_rows:
        ws_sum = wb.create_sheet(title="خلاصه کل", index=0)
        ws_sum.merge_cells("A1:H1")
        ws_sum["A1"] = "📋 خلاصه مقایسه قیمت — همه کانسپت‌ها"
        ws_sum["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
        ws_sum["A1"].fill = PatternFill("solid", start_color="1F3864")
        ws_sum["A1"].alignment = CENTER
        ws_sum.row_dimensions[1].height = 30

        headers = ["کانسپت", "آیتم من", "قیمت من (تومان)", "آیتم رقیب", "قیمت رقیب (تومان)", "رقیب", "اختلاف (تومان)", "وضعیت"]
        for col, h in enumerate(headers, 1):
            cell = ws_sum.cell(row=2, column=col, value=h)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = CENTER
            cell.border = border
        ws_sum.row_dimensions[2].height = 22

        for r, item in enumerate(summary_rows, 3):
            fill = ALT_FILL if r % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            vals = [item["کانسپت"], item["آیتم_من"], item["قیمت_من"], item["آیتم_رقیب"], item["قیمت_رقیب"], item["رقیب"], item["اختلاف"], item["وضعیت"]]
            for col, v in enumerate(vals, 1):
                cell = ws_sum.cell(row=r, column=col, value=v)
                cell.font = NORMAL_FONT
                cell.fill = fill
                cell.alignment = RIGHT if col in [3, 5, 7] else Alignment(horizontal="right", vertical="center", wrap_text=True)
                cell.border = border

        col_widths = [20, 35, 18, 35, 20, 22, 20, 14]
        for i, w in enumerate(col_widths, 1):
            ws_sum.column_dimensions[get_column_letter(i)].width = w
        ws_sum.sheet_view.rightToLeft = True

    return wb

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    import datetime
    print("🚀 MenuRadar starting...")

    my_menu = load_my_menu()
    print(f"✓ My menu loaded: {len(my_menu)} items across {my_menu['کانسپت'].nunique()} concepts")

    competitors = load_competitors()
    print(f"✓ Competitors loaded: {len(competitors)} concept groups")

    # Scrape all competitors
    scraped = {}
    for sheet_name, comp_df in competitors.items():
        print(f"\n📍 Scraping concept: {sheet_name}")
        all_items = []
        for _, row in comp_df.iterrows():
            items = scrape_brand(row["URL"], row["نام_برند"])
            all_items.extend(items)
        scraped[f"_items_{sheet_name}"] = all_items
        print(f"  Total items for {sheet_name}: {len(all_items)}")

    # Merge scraped into competitors dict for build_output
    full_data = {**competitors, **scraped}

    print("\n📊 Building output Excel...")
    wb = build_output(my_menu, full_data, CONCEPT_MAP)

    os.makedirs("reports", exist_ok=True)
    date_str = datetime.date.today().strftime("%Y-%m-%d")
    output_path = f"reports/MenuRadar_{date_str}.xlsx"
    wb.save(output_path)
    print(f"✓ Excel saved: {output_path}")
    return output_path

if __name__ == "__main__":
    main()
